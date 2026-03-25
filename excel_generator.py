"""Longlist — Excel generation with openpyxl, German formatting."""
from __future__ import annotations

import logging
import os
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("longlist.excel")

# Styles
HEADER_FILL = PatternFill(start_color="B8D4E8", end_color="B8D4E8", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="1A1A1A")
CELL_FONT = Font(name="Calibri", size=10, color="333333")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

# Column definitions per package tier
# BASE_COLUMNS maps to the Details endpoint (10 credits) — all packages get these
BASE_COLUMNS = [
    ("Nr.", 5),
    ("Firma", 35),
    ("Rechtsform", 12),
    ("Handelsregister", 22),
    ("Adresse", 35),
    ("PLZ", 8),
    ("Stadt", 18),
    ("Geschäftsführer", 30),
    ("Website", 30),
    ("Telefon", 20),
    ("E-Mail", 30),
    ("Stammkapital", 18),
    ("Branche (WZ-Code)", 35),
    ("Unternehmensgegenstand", 40),
    ("Status", 12),
    ("Gründungsdatum", 16),
]

FINANCIAL_COLUMNS = [
    ("Umsatz (EUR)", 16),
    ("Bilanzsumme (EUR)", 18),
    ("Eigenkapital (EUR)", 18),
    ("Mitarbeiter", 12),
    ("Geschäftsjahr", 14),
]

OWNER_COLUMNS = [
    ("Gesellschafter", 35),
    ("Beteiligung (%)", 15),
]

UBO_COLUMNS = [
    ("Wirtsch. Berechtigte (UBOs)", 35),
]

HOLDINGS_COLUMNS = [
    ("Beteiligungen / Töchter", 40),
]

EMAIL_COLUMNS = [
    ("GF-Email", 30),
]

DOCUMENT_COLUMNS = [
    ("Dokumente (Links)", 50),
    ("Dokumente (URLs)", 60),
]


def _fmt_eur(value: Any) -> str:
    """Format a number as German EUR string (Punkt = Tausender)."""
    if value is None or value == "":
        return ""
    try:
        num = float(value)
        # Convert from cents if value seems to be in cents (> 100000)
        if num > 100_000 and isinstance(value, (int, float)):
            num = num / 100
        if num == int(num):
            return f"{int(num):,}".replace(",", ".")
        return f"{num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return str(value)


def _safe(val: Any, default: str = "") -> str:
    """Return val as human-readable string, flattening dicts/lists."""
    if val is None or val == "" or val == []:
        return default
    if isinstance(val, dict):
        # Try common field names first
        for key in ("name", "value", "text", "description", "url"):
            if key in val and val[key]:
                return str(val[key])
        # Fallback: join non-empty key-value pairs
        parts = [f"{v}" for v in val.values() if v and v is not True]
        return ", ".join(parts) if parts else default
    if isinstance(val, list):
        return ", ".join(_safe(item) for item in val if item)
    return str(val)


def _extract_name(details: dict, company: dict) -> str:
    """Extract company name — details.name can be a dict or string."""
    name = details.get("name") or company.get("name")
    if isinstance(name, dict):
        return name.get("name") or _safe(name)
    return _safe(name)


def _extract_legal_form(details: dict) -> str:
    """Extract legal form, handling both string and dict 'name' field."""
    lf = details.get("legal_form")
    if lf:
        return str(lf).upper()
    name = details.get("name")
    if isinstance(name, dict):
        lf = name.get("legal_form")
        return str(lf).upper() if lf else ""
    return ""


def _extract_register(details: dict, company: dict) -> str:
    """Format register info: 'HRB 232360, AG Berlin (Charlottenburg)'."""
    reg = details.get("company_register") or details.get("register") or {}
    if isinstance(reg, dict):
        rtype = reg.get("register_type", "")
        rnum = reg.get("register_number", "")
        court = reg.get("register_court", "")
        if rtype and rnum:
            parts = [f"{rtype} {rnum}"]
            if court:
                parts.append(f"AG {court}")
            return ", ".join(parts)
    cid = company.get("company_id") or details.get("id") or ""
    return _safe(cid)


def _extract_representatives(details: dict) -> str:
    """Extract Geschäftsführer names from details.representation."""
    reps = details.get("representation") or details.get("representatives") or []
    if isinstance(reps, list):
        names = []
        for r in reps:
            if isinstance(r, dict):
                name = r.get("name") or f"{r.get('first_name', '')} {r.get('last_name', '')}".strip()
                role = r.get("role", "")
                if name:
                    names.append(name)
            elif isinstance(r, str):
                names.append(r)
        return ", ".join(names) if names else ""
    return _safe(reps)


def _extract_address(details: dict) -> tuple[str, str, str]:
    """Extract (street, zip, city) from details.address."""
    addr = details.get("address") or {}
    if isinstance(addr, dict):
        street = addr.get("street") or addr.get("formatted_value") or ""
        plz = addr.get("postal_code") or addr.get("zip_code") or ""
        city = addr.get("city") or ""
        return _safe(street) if street else "", _safe(plz), _safe(city)
    return "", "", ""


def _extract_contact_from_details(details: dict) -> tuple[str, str, str]:
    """Extract (website, phone, email) from details.contact sub-object."""
    contact = details.get("contact") or {}
    if isinstance(contact, dict):
        website = contact.get("website_url") or contact.get("website") or ""
        phone = contact.get("phone") or ""
        email = contact.get("email") or ""
        return _safe(website), _safe(phone), _safe(email)
    return "", "", ""


def _extract_capital(details: dict) -> str:
    """Extract share capital (Stammkapital) from details.capital."""
    capital = details.get("capital") or {}
    if isinstance(capital, dict):
        amount = capital.get("amount") or capital.get("value") or ""
        currency = capital.get("currency") or "EUR"
        if amount:
            try:
                num = float(amount)
                formatted = f"{int(num):,}".replace(",", ".") if num == int(num) else f"{num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                return f"{formatted} {currency}"
            except (ValueError, TypeError):
                return str(amount)
    elif capital and not isinstance(capital, dict):
        return _safe(capital)
    return ""


def _extract_industry_codes(details: dict) -> str:
    """Extract industry/WZ codes from details.industry_codes.wz2025."""
    ic = details.get("industry_codes") or {}
    # API returns nested: {"wz2025": [{code: "62.01"}], "WZ2025": [...]}
    if isinstance(ic, dict):
        codes = ic.get("wz2025") or ic.get("WZ2025") or []
    elif isinstance(ic, list):
        codes = ic
    else:
        return _safe(ic)
    entries = []
    for c in codes:
        if isinstance(c, dict):
            code = c.get("code") or c.get("wz_code") or ""
            desc = c.get("description") or c.get("label") or ""
            if code and desc:
                entries.append(f"{code} — {desc}")
            elif code:
                entries.append(str(code))
        elif isinstance(c, str):
            entries.append(c)
    return ", ".join(entries) if entries else ""


def _extract_purposes(details: dict) -> str:
    """Extract company purpose (Unternehmensgegenstand) from details.purposes."""
    purposes = details.get("purposes") or details.get("purpose") or []
    if isinstance(purposes, list):
        texts = []
        for p in purposes:
            if isinstance(p, dict):
                texts.append(p.get("purpose") or p.get("text") or p.get("description") or "")
            elif isinstance(p, str):
                texts.append(p)
        return "; ".join(texts) if texts else ""
    return _safe(purposes)


def _extract_financials_from_indicators(details: dict) -> dict:
    """Extract financial summary from details.indicators list (most recent year)."""
    indicators = details.get("indicators") or []
    if not isinstance(indicators, list) or not indicators:
        return {}
    latest = indicators[0]
    if isinstance(latest, dict):
        return {
            "revenue": latest.get("revenue"),
            "balance_sheet_total": latest.get("balance_sheet_total"),
            "equity": latest.get("equity"),
            "employees": latest.get("employees"),
            "date": latest.get("date"),
        }
    return {}


def _extract_owners(owners_data: dict) -> list[tuple[str, str]]:
    """Extract list of (name, share%) from owners data."""
    owners_list = owners_data.get("owners") or owners_data.get("shareholders") or []
    if not isinstance(owners_list, list):
        return []
    results = []
    for o in owners_list:
        if isinstance(o, dict):
            name = o.get("name") or ""
            if not name:
                lp = o.get("legal_person") or {}
                np = o.get("natural_person") or {}
                name = lp.get("name") if isinstance(lp, dict) else ""
                if not name and isinstance(np, dict):
                    name = f"{np.get('first_name', '')} {np.get('last_name', '')}".strip()
            share = o.get("percentage_share") or o.get("share_percent") or o.get("share") or ""
            if share and isinstance(share, (int, float)):
                share = f"{share:.1f}"
            results.append((name or "", _safe(share)))
    return results


def _extract_ubos(ubos_data: dict) -> str:
    """Extract UBO names from UBOs data."""
    ubos_list = ubos_data.get("ubos") or ubos_data.get("beneficial_owners") or []
    if not isinstance(ubos_list, list):
        return ""
    names = []
    for u in ubos_list:
        if isinstance(u, dict):
            name = u.get("name") or f"{u.get('first_name', '')} {u.get('last_name', '')}".strip()
            share = u.get("share_percent") or u.get("share") or ""
            if name:
                entry = f"{name} ({share}%)" if share else name
                names.append(entry)
        elif isinstance(u, str):
            names.append(u)
    return ", ".join(names) if names else ""


def _extract_holdings(holdings_data: dict) -> str:
    """Extract holdings/subsidiaries from holdings data."""
    holdings_list = holdings_data.get("holdings") or holdings_data.get("subsidiaries") or []
    if not isinstance(holdings_list, list):
        return ""
    entries = []
    for h in holdings_list:
        if isinstance(h, dict):
            name = h.get("name") or h.get("company_name") or ""
            share = h.get("share_percent") or h.get("share") or ""
            if name:
                entry = f"{name} ({share}%)" if share else name
                entries.append(entry)
        elif isinstance(h, str):
            entries.append(h)
    return ", ".join(entries) if entries else ""


def _extract_documents(details: dict, app_url: str = "") -> str:
    """Extract document info with proxy URLs from details.documents list."""
    docs = details.get("documents") or []
    if not isinstance(docs, list):
        return ""
    entries = []
    for d in docs:
        if isinstance(d, dict):
            doc_id = d.get("id", "")
            title = d.get("name") or d.get("title") or _doc_type_label(d.get("type")) or "Dokument"
            date = d.get("date") or ""
            entry = title
            if date:
                entry += f" ({date})"
            if doc_id and app_url:
                entry += f": {app_url}/doc/{doc_id}"
            entries.append(entry)
        elif isinstance(d, str):
            entries.append(d)
    return "\n".join(entries) if entries else ""


def _extract_document_urls(details: dict, app_url: str = "") -> str:
    """Extract proxy URLs for documents, semicolon-separated."""
    docs = details.get("documents") or []
    if not isinstance(docs, list):
        return ""
    urls = []
    for d in docs:
        if isinstance(d, dict):
            doc_id = d.get("id", "")
            if doc_id and app_url:
                urls.append(f"{app_url}/doc/{doc_id}")
    return "; ".join(urls) if urls else ""


def _doc_type_label(doc_type: str | None) -> str:
    """German label for OpenRegister document types."""
    labels = {
        "articles_of_association": "Gesellschaftsvertrag",
        "current_printout": "Aktueller Abdruck",
        "chronological_printout": "Chronologischer Abdruck",
        "historical_printout": "Historischer Abdruck",
        "structured_information": "Strukturierte Daten",
        "shareholder_list": "Gesellschafterliste",
    }
    return labels.get(doc_type or "", doc_type or "")


def generate_excel(
    companies: list[dict[str, Any]],
    package: str,
    job_id: str,
    output_dir: str = "/tmp",
    app_url: str = "",
) -> str:
    """
    Generate a formatted Excel file from enriched company data.

    Returns the file path.
    """
    columns, includes_financials, includes_owners, includes_ubos, includes_holdings, includes_email = \
        get_columns_for_package(package)

    wb = Workbook()
    ws = wb.active
    ws.title = "Longlist"

    write_company_sheet(ws, companies, columns, package, app_url,
                        includes_financials, includes_owners, includes_ubos,
                        includes_holdings, includes_email)

    # Auto-filter
    if companies:
        last_col = get_column_letter(len(columns))
        last_row = len(companies) + 1
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # Save
    filename = f"Longlist_{job_id}_{package.upper()}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    logger.info("Excel generated: %s (%d companies)", filepath, len(companies))

    return filepath


def get_columns_for_package(package: str) -> tuple[list, bool, bool, bool, bool, bool]:
    """Return (columns, includes_financials, includes_owners, includes_ubos, includes_holdings, includes_email)."""
    from config import PACKAGES
    pkg = PACKAGES.get(package, {})
    inc_fin = pkg.get("includes_financials", False)
    inc_own = pkg.get("includes_owners", False)
    inc_ubo = pkg.get("includes_ubos", False)
    inc_hld = pkg.get("includes_holdings", False)
    inc_email = pkg.get("includes_email_lookup", False)

    columns = list(BASE_COLUMNS)
    if inc_fin:
        columns.extend(FINANCIAL_COLUMNS)
    if inc_own:
        columns.extend(OWNER_COLUMNS)
    if inc_ubo:
        columns.extend(UBO_COLUMNS)
    if inc_hld:
        columns.extend(HOLDINGS_COLUMNS)
    if inc_email:
        columns.extend(EMAIL_COLUMNS)
    columns.extend(DOCUMENT_COLUMNS)

    return columns, inc_fin, inc_own, inc_ubo, inc_hld, inc_email


def write_company_sheet(
    ws,
    companies: list[dict],
    columns: list[tuple[str, int]],
    package: str,
    app_url: str = "",
    includes_financials: bool = False,
    includes_owners: bool = False,
    includes_ubos: bool = False,
    includes_holdings: bool = False,
    includes_email: bool = False,
) -> None:
    """Write header row + company data rows to a worksheet. Shared by all Excel generators."""
    # Write header row
    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Write company rows
    for idx, company in enumerate(companies, 1):
        details = company.get("details", {})
        financials = company.get("financials", {})
        owners = company.get("owners", {})
        ubos = company.get("ubos", {})
        holdings = company.get("holdings", {})

        # Handle error responses
        if "error" in details:
            details = {}
        if "error" in financials:
            financials = {}
        if "error" in owners:
            owners = {}
        if "error" in ubos:
            ubos = {}
        if "error" in holdings:
            holdings = {}

        street, plz, city = _extract_address(details)
        gf = _extract_representatives(details)
        website, phone, company_email = _extract_contact_from_details(details)

        row = idx + 1

        # Base columns (all from Details endpoint)
        values = [
            idx,  # Nr.
            _extract_name(details, company),
            _extract_legal_form(details),
            _extract_register(details, company),
            street,
            plz,
            city,
            gf,
            website,
            phone,
            company_email,
            _extract_capital(details),
            _extract_industry_codes(details),
            _extract_purposes(details),
            _safe(details.get("status")),
            _safe(details.get("incorporated_at")),
        ]

        if includes_financials:
            # Primary: use details.indicators for summary financials
            ind = _extract_financials_from_indicators(details)
            # Fallback: try financials endpoint data
            fin = financials.get("financials") or financials
            annual = fin.get("annual_reports") or fin.get("reports") or []
            latest_report = annual[0] if isinstance(annual, list) and annual else {}

            values.extend([
                _fmt_eur(ind.get("revenue") or latest_report.get("revenue")),
                _fmt_eur(ind.get("balance_sheet_total") or latest_report.get("balance_sheet_total")),
                _fmt_eur(ind.get("equity") or latest_report.get("equity")),
                _safe(ind.get("employees") or latest_report.get("employees")),
                _safe(ind.get("date") or latest_report.get("fiscal_year") or latest_report.get("year")),
            ])

        if includes_owners:
            owner_list = _extract_owners(owners)
            if owner_list:
                values.append(", ".join(n for n, _ in owner_list))
                values.append(", ".join(s for _, s in owner_list))
            else:
                values.extend(["", ""])

        if includes_ubos:
            values.append(_extract_ubos(ubos))

        if includes_holdings:
            values.append(_extract_holdings(holdings))

        if includes_email:
            values.append(_safe(company.get("gf_email")))

        # Document links (always included) — two columns: hyperlinks + raw URLs
        values.append(_extract_documents(details, app_url=app_url))
        values.append(_extract_document_urls(details, app_url=app_url))

        # Write values
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = CELL_FONT
            cell.border = THIN_BORDER
            if col_idx == 1:  # Nr. column centered
                cell.alignment = Alignment(horizontal="center")
