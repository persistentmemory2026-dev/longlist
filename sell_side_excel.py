"""Longlist — Multi-tab Excel generation for sell-side buyer longlists."""
import logging
import os
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from excel_generator import (
    HEADER_FILL,
    HEADER_FONT,
    THIN_BORDER,
    get_columns_for_package,
    write_company_sheet,
)

logger = logging.getLogger("longlist.sell_side_excel")


def generate_sell_side_excel(
    buyer_groups: list[dict[str, Any]],
    package: str,
    job_id: str,
    target_name: str = "",
    output_dir: str = "/tmp",
    app_url: str = "",
) -> str:
    """
    Generate a multi-tab Excel for sell-side buyer longlists.

    Tab 1: Übersicht (summary of all buyer groups)
    Tab 2-N: One tab per buyer group with enriched company data
    """
    columns, inc_fin, inc_own, inc_ubo, inc_hld, inc_email = get_columns_for_package(package)

    wb = Workbook()

    # Tab 1: Übersicht
    ws_overview = wb.active
    ws_overview.title = "Übersicht"

    # Title
    title_cell = ws_overview.cell(row=1, column=1, value=f"Buyer-Longlist: {target_name}")
    title_cell.font = Font(name="Georgia", size=16, bold=True, color="1b4332")
    ws_overview.merge_cells("A1:D1")

    # Subtitle
    total_companies = sum(len(g.get("companies", [])) for g in buyer_groups)
    ws_overview.cell(row=2, column=1, value=f"{total_companies} Unternehmen in {len(buyer_groups)} Käufergruppen")
    ws_overview.cell(row=2, column=1).font = Font(name="Calibri", size=11, color="71717a")

    # Header row
    headers = [("Käufergruppe", 30), ("Beschreibung", 45), ("Begründung", 45), ("Anzahl", 10)]
    for col_idx, (name, width) in enumerate(headers, 1):
        cell = ws_overview.cell(row=4, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws_overview.column_dimensions[get_column_letter(col_idx)].width = width

    # Group rows
    for i, group in enumerate(buyer_groups):
        row = 5 + i
        companies = group.get("companies", [])
        ws_overview.cell(row=row, column=1, value=group.get("name", "")).font = Font(name="Calibri", size=11, bold=True)
        ws_overview.cell(row=row, column=2, value=group.get("description", ""))
        ws_overview.cell(row=row, column=3, value=group.get("rationale", ""))
        count_cell = ws_overview.cell(row=row, column=4, value=len(companies))
        count_cell.alignment = Alignment(horizontal="center")
        for col_idx in range(1, 5):
            ws_overview.cell(row=row, column=col_idx).border = THIN_BORDER

    ws_overview.freeze_panes = "A5"

    # Tab 2-N: One tab per buyer group
    for group in buyer_groups:
        companies = group.get("companies", [])
        if not companies:
            continue

        # Excel tab names max 31 chars, no special chars
        tab_name = group.get("name", "Gruppe")[:31]
        ws = wb.create_sheet(title=tab_name)

        write_company_sheet(
            ws, companies, columns, package, app_url,
            inc_fin, inc_own, inc_ubo, inc_hld, inc_email,
        )

        # Auto-filter
        if companies:
            last_col = get_column_letter(len(columns))
            last_row = len(companies) + 1
            ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # Save
    filename = f"Longlist_{job_id}_SELL_SIDE.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    logger.info("Sell-side Excel generated: %s (%d companies, %d groups)",
                filepath, total_companies, len(buyer_groups))

    return filepath
